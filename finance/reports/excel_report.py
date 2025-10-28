# finance/reports/excel_report.py - ИСПРАВИТЬ
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Sum
import io


def generate_excel_report(user, start_date, end_date, categories=None):
    """Генерация Excel отчета с графиками"""
    from finance.models import Transaction

    # Получаем транзакции
    transactions = Transaction.objects.filter(
        user=user,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')

    if categories:
        transactions = transactions.filter(category__name__in=categories)

    # Создание workbook
    wb = openpyxl.Workbook()

    # === Лист с детальными данными ===
    ws_data = wb.active
    ws_data.title = "Детальные данные"

    # Заголовки
    headers = ['Дата', 'Тип', 'Категория', 'Сумма', 'Описание']
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Данные транзакций
    for row, transaction in enumerate(transactions, 2):
        ws_data.cell(row=row, column=1, value=transaction.date.strftime('%d.%m.%Y'))
        ws_data.cell(row=row, column=2, value='Доход' if transaction.type == 'income' else 'Расход')
        ws_data.cell(row=row, column=3, value=transaction.category.name)
        ws_data.cell(row=row, column=4, value=float(transaction.amount))  # Конвертируем Decimal в float
        ws_data.cell(row=row, column=5, value=transaction.description or '')

    # Автоширина колонок
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width

    # === Лист со сводкой ===
    ws_summary = wb.create_sheet("Сводка")

    # Основные показатели - КОНВЕРТИРУЕМ Decimal в float
    total_income_result = transactions.filter(type='income').aggregate(Sum('amount'))['amount__sum']
    total_income = float(total_income_result) if total_income_result else 0.0

    total_expense_result = transactions.filter(type='expense').aggregate(Sum('amount'))['amount__sum']
    total_expense = float(total_expense_result) if total_expense_result else 0.0

    balance = total_income - total_expense

    summary_data = [
        ['Показатель', 'Значение'],
        ['Период отчета', f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"],
        ['Общее количество операций', len(transactions)],
        ['Количество доходов', transactions.filter(type='income').count()],
        ['Количество расходов', transactions.filter(type='expense').count()],
        ['Общие доходы', f"{total_income:.2f} руб."],
        ['Общие расходы', f"{total_expense:.2f} руб."],
        ['Баланс', f"{balance:.2f} руб."],
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)

    # === Лист с анализом по категориям ===
    ws_categories = wb.create_sheet("Анализ по категориям")

    # Анализ расходов по категориям - КОНВЕРТИРУЕМ Decimal в float
    expense_categories = {}
    for transaction in transactions.filter(type='expense'):
        cat_name = transaction.category.name
        if cat_name not in expense_categories:
            expense_categories[cat_name] = 0.0
        expense_categories[cat_name] += float(transaction.amount)  # Конвертируем здесь

    # Сортируем по убыванию суммы
    sorted_categories = sorted(expense_categories.items(), key=lambda x: x[1], reverse=True)

    ws_categories.cell(row=1, column=1, value="Категория").font = Font(bold=True)
    ws_categories.cell(row=1, column=2, value="Сумма расходов").font = Font(bold=True)
    ws_categories.cell(row=1, column=3, value="Доля").font = Font(bold=True)

    for row, (category, amount) in enumerate(sorted_categories, 2):
        # ИСПРАВЛЕННАЯ СТРОКА: используем float значения
        percentage = (amount / total_expense * 100) if total_expense > 0 else 0
        ws_categories.cell(row=row, column=1, value=category)
        ws_categories.cell(row=row, column=2, value=amount)
        ws_categories.cell(row=row, column=3, value=f"{percentage:.1f}%")

    # Убираем круговую диаграмму пока что (чтобы избежать ошибок)
    # if sorted_categories:
    #     pie_chart = PieChart()
    #     labels = Reference(ws_categories, min_col=1, min_row=2, max_row=len(sorted_categories)+1)
    #     data = Reference(ws_categories, min_col=2, min_row=1, max_row=len(sorted_categories)+1)
    #     pie_chart.add_data(data, titles_from_data=True)
    #     pie_chart.set_categories(labels)
    #     pie_chart.title = "Расходы по категориям"
    #     ws_categories.add_chart(pie_chart, "E2")

    # Сохранение в buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer